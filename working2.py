from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook

# Initialize the webdriver
driver = webdriver.Chrome("E:\Code\Web Scraping\chromedriver.exe")

# Open the webpage
driver.get("https://www.fact.co.in/")

# Wait for the page to load completely
driver.implicitly_wait(10)

# Get the column names from the user
text_column_name = input("Enter the column name for the text: ")
link_column_name = input("Enter the column name for the link: ")

# Create empty lists for the text and links
text_list = []
link_list = []

# Find all the links on the webpage
links = driver.find_elements(By.XPATH, "//a")

# Loop through each link and extract the text and link URL
for link in links:
    text = link.text
    href = link.get_attribute("href")

    # Append the text and link URL to their respective lists
    text_list.append(text)
    link_list.append(href)

# Combine the text and link lists into a dictionary
data = {text_column_name: text_list, link_column_name  : link_list}

# Create a new workbook and worksheet
wb = Workbook()
ws = wb.active

# Write the column names to the first row of the worksheet
ws.append([text_column_name, link_column_name ])

# Write the data to the worksheet
for row in range(len(text_list)):
    ws.cell(row=row+2, column=1, value=text_list[row])
    ws.cell(row=row+2, column=2, value=link_list[row])

# Save the workbook
wb.save("fact_scraped.xlsx")

# Close the webdriver
driver.quit()
